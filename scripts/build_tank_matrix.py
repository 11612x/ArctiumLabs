"""Extract Shell pre-cargo matrix from Excel export to JS data file."""
import json
import re
import hashlib
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = Path.home() / "OneDrive" / "Desktop" / "Shell Ship_PreCargo_Matrix 2024.xlsx"
OUT_DATA = ROOT / "tank-cleaning-data.js"


def load_sheet_rows(name: str, max_col: int) -> list[list[str]]:
    import openpyxl

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[name]
    rows = []
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col, values_only=True):
        rows.append([str(c).strip() if c is not None else "" for c in r])
    wb.close()
    return rows

# Display labels for dropdowns (order matches matrix where applicable)
PREVIOUS_UI = [
    ("p01", "ULSD 10ppm", "Up to 50 ppm sulphur ULSD"),
    ("p02", "ULSD 50ppm", "Up to 50 ppm sulphur ULSD"),
    ("p00", "Diesels/Gasoils up to 15% FAME / Biodiesel", "up to 15% FAME / Biodiesel"),
    ("p03", "V-Power Diesel (Detergent Additivated)", "Detergent addivated 10ppm"),
    ("p04", "GTL Kero/Diesel/n-Paraffins/HVO/HDRD", "GTL kero/diesel/paraffin, HVO, HDRD"),
    ("p05", "Dyed Gasoil 500/2000ppm", "DYED gas oil/IGO"),
    ("p06", "Undyed Gasoil 500/2000ppm", "UNDYED gas oil/IGO"),
    ("p07", "Gasoline 10/50ppm", "Up to 500 ppm sulphur gasoline"),
    ("p08", "Dyed Gasoline", "Dyed gasoline"),
    ("p09", "Kerosene/Burning Oil", "Jet A1, AVTUR & undyed kerosene"),
    ("p10", "Synthetic Aviation Fuel / HEFA-SPK (SAF)", "HEFA-SPK"),
    ("p11", "Jet A1 / AVTUR / DPK", "Jet A1, AVTUR & undyed kerosene"),
    ("p12", "AVGAS 100LL", "AVGAS 100LL"),
    ("p13", "MTBE & ETBE", "MTBE & ETBE"),
    ("p14", "Xylene / Toluene / Mixed Aromatics", "Mixed Aromatics"),
    ("p15", "Ethyl Benzene", "Ethyl Benzene"),
    ("p16", "Benzene", "UN 1114 \nBenzene"),
    ("p17", "Benzene Heart Cut / BHC", "Benzene Heart Cut, Pygas"),
    ("p18", "Natural Gas Condensate C5+", "Natural gas condensate C5+"),
    ("p19", "Pygas / Pyrolysis Gasoline", "Pyrolysis gasoline"),
    ("p20", "Cat Cracked Gasoline / LCCG / FRCCG", "Cat cracked gasoline"),
    ("p21", "Platformate / Reformate", "Platformate, Reformate"),
    ("p22", "Naphtha / bio-Naphtha / Platfeed", "Naphtha, bio-Naphtha"),
    ("p23", "Isomerate / Alkylate", "Isomerate, Akylate"),
    ("p24", "FAME / Distillates >15% FAME", ">15% FAME content"),
    ("p25", "Ethanol / Methanol", "Ethanol"),
    ("p26", "Cycle Oils / LCO / HCO", "Cycle oils, LCO & HCO"),
    ("p27", "Vegetable Oils", "Vegetable oils"),
    ("p28", "GTL Base Oils / Shell XHVI", "GTL Base Oils"),
    ("p29", "Lube Base Oil", "Lube Base Oil"),
    ("p30", "Black Oils / Crude / Fuel Oil / Dirty Condensate", "Black oils incl"),
]

TO_LOAD_UI = [
    ("l01", "ULSD 10ppm", "ULSD - 10 ppm"),
    ("l02", "ULSD 50ppm", "ULSD - 50 ppm"),
    ("l03", "V-Power Diesel (Detergent Additivated)", "Detergent Addivated 10ppm"),
    ("l04", "GTL Kero/Diesel/n-Paraffins/HVO/HDRD", "GTL kero/diesel/n-paraffins, HVO, HDRD"),
    ("l05", "Dyed Gasoil 500/2000ppm", "DYED & UNDYED gas oil"),
    ("l06", "Undyed Gasoil 500/2000ppm", "DYED & UNDYED gas oil"),
    ("l07", "Gasoline 10ppm", "10ppm Gasoline"),
    ("l08", "Gasoline 50ppm", "up to 500ppm S gasoline"),
    ("l09", "Kerosene/Burning Oil", "Kerosene, Burning oil"),
    ("l10", "Synthetic Aviation Fuel / HEFA-SPK (SAF)", "HEFA-SPK", True),
    ("l11", "Jet A1 / AVTUR / DPK", "Jet A1, AVTUR", True),
    ("l12", "AVGAS 100LL", "AVGAS", True),
    ("l13", "MTBE & ETBE", "MTBE & ETBE"),
    ("l14", "Xylene / Toluene / Mixed Aromatics", "Mixed Aromatics"),
    ("l15", "Ethyl Benzene", "Ethyl Benzene"),
    ("l16", "Benzene", "UN 1114 \nBenzene"),
    ("l17", "Benzene Heart Cut / BHC", "Benzene Heart Cut, BHC"),
    ("l18", "Natural Gas Condensate C5+", "Natural Gas Condensate"),
    ("l19", "Pygas / Pyrolysis Gasoline", "Pyrolysis Gasoline"),
    ("l20", "Cat Cracked Gasoline / LCCG / FRCCG", "Cat Cracked Gasoline"),
    ("l21", "Platformate / Reformate", "Platformate, Reformate"),
    ("l22", "Naphtha / bio-Naphtha / Platfeed", "Naphtha, bio-Naphtha"),
    ("l23", "Isomerate / Alkylate", "Isomerate/Alkylate"),
    ("l24", "FAME / Distillates >15% FAME", ">15% FAME content"),
    ("l25", "Ethanol / Methanol", "Ethanol"),
    ("l26", "Cycle Oils / LCO / HCO", "Cycle oils, HCO, LCO"),
    ("l27", "GTL Base Oils / Shell XHVI", "GTL base oils/Shell XHVI"),
    ("l28", "Lube Base Oil", "Lube base oils/GTL base oils"),
]

AVIATION_LOAD_IDS = {"l10", "l11", "l12"}
L3C_CRITICAL_PREV_IDS = {"p24", "p27", "p29"}  # FAME, veg oils, lube

# Wording overrides (matrix source may name a specific company; app stays charterer-neutral)
DEFINITION_OVERRIDES = {
    "Seek Guidance": (
        "You should seek further guidance from the charterer as more stringent tank "
        "preparations and analytical checks may be required to prepare tanks satisfactory."
    ),
}


def norm(s: str) -> str:
    return re.sub(r"\s+", " ", s.replace("\n", " ").strip()).lower()


def find_header(headers: list[str], match: str) -> str:
    m = norm(match)
    hits = [h for h in headers if m in norm(h)]
    if not hits:
        raise ValueError(f"No header match for {match!r}")
    return min(hits, key=len)


def main() -> None:
    if not EXCEL_PATH.is_file():
        raise FileNotFoundError(f"Excel not found: {EXCEL_PATH}")

    rows = load_sheet_rows("4 Ship Pre-Cargo Matrix", 32)
    prev_headers = [c for c in rows[2][3:] if c.strip()]

    to_load_headers: list[str] = []
    raw_by_load: dict[str, dict[str, str]] = {}
    for row in rows[3:]:
        name = row[2].strip() if len(row) > 2 else ""
        if not name.startswith("UN"):
            continue
        vals = [v.strip() for v in row[3 : 3 + len(prev_headers)]]
        while len(vals) < len(prev_headers):
            vals.append("")
        to_load_headers.append(name)
        raw_by_load[name] = dict(zip(prev_headers, vals))

    defs_rows = load_sheet_rows("3 Key to Definitions", 5)
    definitions: dict[str, str] = {}
    for row in defs_rows:
        if len(row) < 4:
            continue
        key, detail = row[2].strip(), row[3].strip()
        if not key or not detail:
            continue
        if key in ("Key",) or key.startswith("EXPLANATORY"):
            continue
        if key.startswith("EWD Note") or key.startswith("An alternative"):
            continue
        if key.isdigit() or key.startswith("Refers") or key.startswith("Where a"):
            continue
        if key.startswith("\u0096") or key.startswith("-"):
            continue
        definitions[key] = detail
    definitions.update(DEFINITION_OVERRIDES)

    prev_map = {pid: find_header(prev_headers, m) for pid, _, m in PREVIOUS_UI}
    load_map: dict[str, str] = {}
    for item in TO_LOAD_UI:
        pid, _, m = item[0], item[1], item[2]
        if pid in load_map:
            continue
        load_map[pid] = find_header(to_load_headers, m)

    matrix: dict[str, dict[str, str]] = {}
    for lid, load_hdr in load_map.items():
        matrix[lid] = {}
        for pid, prev_hdr in prev_map.items():
            matrix[lid][pid] = raw_by_load[load_hdr][prev_hdr]

    previous_products = [
        {"id": pid, "label": label, "header": prev_map[pid]}
        for pid, label, _ in PREVIOUS_UI
    ]
    to_load_products = []
    for item in TO_LOAD_UI:
        pid, label = item[0], item[1]
        hdr = load_map[pid]
        to_load_products.append(
            {
                "id": pid,
                "label": label,
                "header": hdr,
                "aviation": pid in AVIATION_LOAD_IDS,
            }
        )

    # Full raw matrix for audit (all Excel cells)
    full_matrix = {lh: raw_by_load[lh] for lh in to_load_headers}

    payload = {
        "meta": {
            "source": "Shell Ship Pre-Cargo Matrix Issue 11.0, April 2024",
            "previousColumns": len(prev_headers),
            "toLoadRows": len(to_load_headers),
            "rawCells": sum(len(v) for v in raw_by_load.values()),
        },
        "previousProducts": previous_products,
        "toLoadProducts": to_load_products,
        "definitions": definitions,
        "matrix": matrix,
        "fullMatrix": full_matrix,
        "previousHeaders": prev_headers,
        "toLoadHeaders": to_load_headers,
        "aviationLoadIds": sorted(AVIATION_LOAD_IDS),
        "l3cCriticalPrevIds": sorted(L3C_CRITICAL_PREV_IDS),
    }

    OUT_DATA.write_text(
        "/* Shell Ship Pre-Cargo Matrix Issue 11.0 — generated from Excel */\n"
        "window.TANK_CLEANING_DATA = "
        + json.dumps(payload, ensure_ascii=False)
        + ";\n",
        encoding="utf-8",
    )

    ui_cells = sum(len(v) for v in matrix.values())
    log_path = ROOT / "_build_log.txt"
    log_path.write_text(
        f"prev_headers={len(prev_headers)} to_load={len(to_load_headers)} "
        f"raw_cells={payload['meta']['rawCells']} ui_cells={ui_cells} "
        f"defs={len(definitions)}\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
